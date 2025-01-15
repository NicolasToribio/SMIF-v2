from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Create a workbook object
#wb = Workbook()



#load existing spreadsheet
wb = load_workbook('SMIF Portfolio Tracker.xlsx', data_only=True)

#Create an active worksheet
wstransaction = wb['Transaction']

# Set a variable
ticker = wstransaction["B6"].value
total_value = wstransaction["G6"].value

#Print something from our spreadsheet
print(f'{ticker}: {total_value}')
#print(wb.sheetnames)



'''from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Create a workbook object
#wb = Workbook()



#load existing spreadsheet
wb = load_workbook('SMIF Portfolio Tracker.xlsx')

#Create an active worksheet
ws = wb.active

# Set a variable
name = ws["A2"].value
color = ws["B2"].value

#Print something from our spreadsheet
print(f'{name}: {color}')
print(wb.sheetnames)'''